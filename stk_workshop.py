"""
Asystent STK — Aplikacja Warsztatowa
Budowanie profili kompetencji dla wielu stanowisk + analiza potrzeb (poziom oczekiwany).
Wynik: JSON (import do Asystenta STK Trenera) + XLSX + DOCX z wykresem radarowym.
"""
from __future__ import annotations

import io
import json
import os

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stk_data import (
    CompanyProfile, CompetencyProfile, NeedsAssessmentItem, NeedsAssessment,
    CriticalIncident,
    LEVEL_LABELS, LEVEL_KEYS, LEVEL_COLORS,
    COMPETENCY_CATALOG, CATEGORY_LABELS,
    task_priority, tasks_to_str,
)
from stk_ai import generate_competency_profile, generate_single_competency


# ---------------------------------------------------------------------------
# Kolory EA
# ---------------------------------------------------------------------------
_NAVY = "1B4F8A"
_LIGHT = "E8F0FB"
_RED   = "FFCCCC"
_AMBER = "FFF0CC"
_GREEN = "CCFFCC"


# ---------------------------------------------------------------------------
# Eksport XLSX (pojedyncze stanowisko — czytelny raport)
# ---------------------------------------------------------------------------
def _hdr(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(wrap_text=True, vertical="top")


def _cell(ws, row: int, col: int, value, bold: bool = False, bg: str | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Montserrat", bold=bold, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="top")


def build_workshop_xlsx(
    profile: CompetencyProfile,
    assessment: NeedsAssessment | None,
) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Profil_kompetencji"
    ws1.merge_cells("A1:I1")
    t = ws1.cell(row=1, column=1,
                 value=f"Profil kompetencji — {profile.company.position_name} ({profile.company.company_name})")
    t.font = Font(name="Montserrat", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 24

    info = [
        ("Firma",             profile.company.company_name),
        ("Branża",            profile.company.industry),
        ("Wielkość",          profile.company.size),
        ("Stanowisko",        profile.company.position_name),
        ("Poziom",            profile.company.position_level),
        ("Kluczowe zadania",  tasks_to_str(profile.company.key_tasks_list) if profile.company.key_tasks_list else profile.company.key_tasks),
        ("Kultura / kontekst", profile.company.culture_notes),
    ]
    for i, (label, val) in enumerate(info, start=2):
        _cell(ws1, i, 1, label, bold=True, bg=_LIGHT)
        ws1.merge_cells(f"B{i}:I{i}")
        _cell(ws1, i, 2, val)

    hr = len(info) + 3
    for col, h in enumerate([
        "Lp.", "Kompetencja", "Kategoria", "Definicja",
        "Poziom 1 — Niedostateczny", "Poziom 2 — Podstawowy", "Poziom 3 — Dobry",
        "Poziom 4 — Bardzo dobry", "Poziom 5 — Wybitny",
    ], 1):
        _hdr(ws1, hr, col, h)

    for i, comp in enumerate(profile.competencies, 1):
        r = hr + i
        bg = _LIGHT if i % 2 == 0 else None
        _cell(ws1, r, 1, i, bg=bg)
        _cell(ws1, r, 2, comp.name, bold=True, bg=bg)
        _cell(ws1, r, 3, CATEGORY_LABELS.get(comp.category, comp.category), bg=bg)
        _cell(ws1, r, 4, comp.definition, bg=bg)
        for j, lk in enumerate(LEVEL_KEYS, 5):
            _cell(ws1, r, j, comp.get_level_description(lk), bg=bg)
        ws1.row_dimensions[r].height = 80

    for col, w in enumerate([4, 22, 14, 36, 28, 28, 28, 28, 28], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    if assessment:
        ws2 = wb.create_sheet("Analiza_potrzeb")
        ws2.merge_cells("A1:F1")
        t2 = ws2.cell(
            row=1, column=1,
            value=f"Analiza potrzeb — {assessment.assessor_name} ({assessment.assessor_role})"
        )
        t2.font = Font(name="Montserrat", bold=True, size=13, color="FFFFFF")
        t2.fill = PatternFill("solid", fgColor=_NAVY)
        t2.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 24

        for col, h in enumerate([
            "Lp.", "Kompetencja", "Poziom oczekiwany", "Ważność (1-5)", "Priorytet (waga × poziom)",
        ], 1):
            _hdr(ws2, 2, col, h)

        sorted_items = sorted(assessment.items, key=lambda x: x.desired_level * x.importance, reverse=True)
        for i, item in enumerate(sorted_items, 1):
            r = i + 2
            bg = _LIGHT if i % 2 == 0 else None
            _cell(ws2, r, 1, i, bg=bg)
            _cell(ws2, r, 2, item.competency_name, bold=True, bg=bg)
            _cell(ws2, r, 3, f"{item.desired_level} — {LEVEL_LABELS[item.desired_level]}", bg=bg)
            _cell(ws2, r, 4, item.importance, bg=bg)
            _cell(ws2, r, 5, item.desired_level * item.importance, bg=bg)

        for col, w in enumerate([4, 30, 26, 12, 18], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Eksport JSON wielu stanowisk
# ---------------------------------------------------------------------------
def build_positions_json(positions: list[dict]) -> str:
    """Eksportuje listę stanowisk do JSON dla trenera (stk_app.py tab6).

    Każde stanowisko może zawierać: profile, needs_assessment, incidents.
    """
    exported = []
    for entry in positions:
        p: CompetencyProfile = entry["profile"]
        na: NeedsAssessment | None = entry.get("needs")
        incs: list[CriticalIncident] = entry.get("incidents", [])
        pos_data: dict = {"profile": p.to_json()}
        if na:
            pos_data["needs_assessment"] = na.to_json()
        if incs:
            pos_data["incidents"] = [i.to_dict() for i in incs]
        exported.append(pos_data)
    return json.dumps({"version": 2, "positions": exported}, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# Resetowanie formularza
# ---------------------------------------------------------------------------
def _reset_form() -> None:
    """Usuwa klucze aktualnie edytowanego stanowiska z session_state."""
    keys_to_clear = [
        "profile", "needs_assessment", "company", "custom_competencies",
        "c_name", "c_industry", "c_size", "c_level", "c_culture",
        "c_position", "c_tasks", "c_extra", "custom_comp_name",
        "assess_name", "incidents", "key_tasks_list",
    ]
    for key in keys_to_clear:
        st.session_state.pop(key, None)
    # Dynamiczne klucze widgetów
    for key in list(st.session_state.keys()):
        if key.startswith(("cat_select_", "des_", "imp_", "rm_comp_", "rm_custom_",
                           "inc_comp_", "inc_sit_", "inc_act_", "inc_actors_",
                           "inc_reas_", "inc_res_", "inc_best_", "inc_worst_",
                           "del_ktask_", "new_ktask_")):
            del st.session_state[key]


# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Profil Kompetencji — Warsztat",
    page_icon="🗺️",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
if "saved_positions" not in st.session_state:
    st.session_state["saved_positions"] = []
if "custom_competencies" not in st.session_state:
    st.session_state["custom_competencies"] = []
if "incidents" not in st.session_state:
    st.session_state["incidents"] = []

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("Profil Kompetencji")
    st.caption("Narzędzie warsztatowe — Enterprise Advisors")

    # API key
    default_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not default_key:
        try:
            default_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            default_key = ""
    if not default_key:
        env_path = os.path.join(os.path.dirname(__file__), ".env")
        if os.path.exists(env_path):
            for line in open(env_path):
                line = line.strip()
                if line.startswith("ANTHROPIC_API_KEY="):
                    default_key = line.split("=", 1)[1].strip().strip('"').strip("'")

    if default_key:
        api_key = default_key
        st.caption("Klucz API: skonfigurowany")
    else:
        api_key = st.text_input("Klucz API Anthropic", type="password")

    st.divider()
    st.markdown(
        "**Jak korzystać:**\n\n"
        "1. **Zakładka 1** — opisz firmę, wybierz kompetencje, wygeneruj opisy\n"
        "2. **Zakładka 2** — ustaw poziomy oczekiwane, zapisz stanowisko\n"
        "3. **Powtórz** dla każdego stanowiska\n"
        "4. **Pobierz pliki** poniżej i wyślij trenerowi\n\n"
        "*Enterprise Advisors*"
    )

    st.divider()

    saved: list[dict] = st.session_state["saved_positions"]

    if saved:
        st.subheader(f"Zapisane stanowiska ({len(saved)})")
        to_del = []
        for idx, entry in enumerate(saved):
            p = entry["profile"]
            col_a, col_b = st.columns([4, 1])
            col_a.markdown(f"**{idx+1}.** {p.company.position_name}")
            col_a.caption(f"{p.company.company_name} · {len(p.competencies)} kompetencji")
            if col_b.button("✕", key=f"del_pos_{idx}"):
                to_del.append(idx)
        for idx in reversed(to_del):
            st.session_state["saved_positions"].pop(idx)
        if to_del:
            st.rerun()

        if st.button("Usuń wszystkie", key="clear_all_positions"):
            st.session_state["saved_positions"] = []
            st.rerun()

        st.divider()
        st.subheader("Pobierz wszystkie stanowiska")

        safe_base = saved[0]["profile"].company.company_name.replace(" ", "_") if saved else "warsztat"

        # JSON — dla trenera
        json_str = build_positions_json(saved)
        st.download_button(
            "Pobierz JSON (dla trenera)",
            data=json_str.encode("utf-8"),
            file_name=f"profile_{safe_base}.json",
            mime="application/json",
            type="primary",
        )

        # XLSX zbiorczy
        try:
            from stk_export import build_combined_xlsx
            profiles_list = [e["profile"] for e in saved]
            xlsx_bytes = build_combined_xlsx(profiles_list)
            st.download_button(
                "Pobierz XLSX (zbiorczy)",
                data=xlsx_bytes,
                file_name=f"profile_{safe_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.caption(f"XLSX niedostępny: {e}")

        # DOCX zbiorczy z wykresami
        try:
            from stk_export import build_combined_docx
            needs_list = [e["needs"] for e in saved]
            docx_bytes = build_combined_docx(profiles_list, needs_list)
            st.download_button(
                "Pobierz DOCX (raport z wykresami)",
                data=docx_bytes,
                file_name=f"Mapa_Kompetencji_{safe_base}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        except Exception as e:
            st.caption(f"DOCX niedostępny: {e}")

    else:
        st.info("Brak zapisanych stanowisk.\nUzupełnij profil i zapisz (zakładka 2).")


# ---------------------------------------------------------------------------
# Zakładki
# ---------------------------------------------------------------------------
tab1, tab2, tab3, tab_osp = st.tabs([
    "1. Profil kompetencji",
    "2. Analiza potrzeb",
    "3. Incydenty krytyczne",
    "OSP",
])


# ===========================================================================
# ZAKŁADKA 1: Profil kompetencji
# ===========================================================================
with tab1:
    st.header("Profiler kompetencji")

    if st.session_state["saved_positions"]:
        st.info(
            f"Zapisano {len(st.session_state['saved_positions'])} stanowisk. "
            "Możesz dodać kolejne lub pobrać wyniki z panelu bocznego."
        )

    st.subheader("Krok 1: Firma i stanowisko")
    col1, col2 = st.columns(2)
    with col1:
        company_name = st.text_input("Nazwa firmy", key="c_name")
        industry = st.text_input("Branża", key="c_industry",
                                 placeholder="np. IT, produkcja, usługi finansowe")
        size = st.selectbox("Wielkość", ["<50", "50-250", ">250"], key="c_size")
        culture = st.text_area("Kultura / kontekst", key="c_culture", height=80,
                               placeholder="np. flat structure, agile, korporacja...")
    with col2:
        position = st.text_input("Nazwa stanowiska", key="c_position")
        level = st.selectbox("Poziom", [
            "specjalista", "starszy specjalista", "kierownik zespołu",
            "kierownik działu", "dyrektor", "zarząd",
        ], key="c_level")
        extra = st.text_area("Dodatkowy kontekst (opcj.)", key="c_extra", height=80)

    # --- Krok 2: Analiza pracy — kluczowe zadania ---
    st.divider()
    st.subheader("Krok 2: Analiza pracy — kluczowe zadania")
    st.caption("Oceń każde zadanie w skali 1-3. Zadania o wysokiej ważności wyznaczają kompetencje kluczowe.")

    if "key_tasks_list" not in st.session_state:
        st.session_state["key_tasks_list"] = []

    _ktl = st.session_state["key_tasks_list"]
    if _ktl:
        _hcols = st.columns([4, 1.2, 1.2, 1.2, 1.5, 0.6])
        for _h, _col in zip(["Zadanie", "Częstość", "Ważność", "Trudność", "Priorytet", ""], _hcols):
            _col.markdown(f"**{_h}**")
        _to_del = []
        for _ti, _task in enumerate(_ktl):
            _rc = st.columns([4, 1.2, 1.2, 1.2, 1.5, 0.6])
            _rc[0].write(_task["name"])
            _rc[1].write(str(_task["frequency"]))
            _rc[2].write(str(_task["importance"]))
            _rc[3].write(str(_task["difficulty"]))
            _prio = task_priority(_task["importance"])
            _pcolor = "#1B4F8A" if _prio == "Wysoki" else ("#e65c00" if _prio == "Sredni" else "#888888")
            _rc[4].markdown(f"<span style='font-weight:bold;color:{_pcolor}'>{_prio}</span>", unsafe_allow_html=True)
            if _rc[5].button("✕", key=f"del_ktask_{_ti}"):
                _to_del.append(_ti)
        if _to_del:
            for _i in reversed(_to_del):
                st.session_state["key_tasks_list"].pop(_i)
            st.rerun()

    with st.expander("+ Dodaj zadanie", expanded=not bool(_ktl)):
        _nc1, _nc2, _nc3, _nc4 = st.columns([4, 1.2, 1.2, 1.2])
        with _nc1:
            _new_name = st.text_input("Nazwa zadania", key="new_ktask_name",
                                      placeholder="np. Rozpoznanie potrzeb i prezentacja oferty")
        with _nc2:
            _new_freq = st.selectbox("Częstość", [1, 2, 3], index=1, key="new_ktask_freq",
                                     help="1=rzadko  2=regularnie  3=często")
        with _nc3:
            _new_imp = st.selectbox("Ważność", [1, 2, 3], index=1, key="new_ktask_imp",
                                    help="1=mała  2=średnia  3=wysoka")
        with _nc4:
            _new_diff = st.selectbox("Trudność", [1, 2, 3], index=1, key="new_ktask_diff",
                                     help="1=prosta  2=umiarkowana  3=złożona")
        if st.button("Dodaj zadanie", key="btn_add_ktask"):
            if not _new_name.strip():
                st.warning("Podaj nazwę zadania.")
            else:
                st.session_state["key_tasks_list"].append({
                    "name": _new_name.strip(),
                    "frequency": _new_freq,
                    "importance": _new_imp,
                    "difficulty": _new_diff,
                })
                for _k in ["new_ktask_name", "new_ktask_freq", "new_ktask_imp", "new_ktask_diff"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    st.divider()
    st.subheader("Krok 3: Wybierz kompetencje z katalogu (max 10)")

    selected_from_catalog: dict[str, list[str]] = {}
    cat_cols = st.columns(4)
    for idx, (cat_key, cat_label) in enumerate(CATEGORY_LABELS.items()):
        with cat_cols[idx]:
            st.markdown(f"**{cat_label}**")
            selected_from_catalog[cat_key] = st.multiselect(
                f"Wybierz {cat_label.lower()}",
                COMPETENCY_CATALOG[cat_key],
                key=f"cat_select_{cat_key}",
                label_visibility="collapsed",
            )

    all_selected: list[str] = []
    for names in selected_from_catalog.values():
        all_selected.extend(names)

    st.divider()
    col_add1, col_add2 = st.columns([3, 1])
    with col_add1:
        custom_name = st.text_input(
            "Dopisz własną kompetencję (nazwa)",
            key="custom_comp_name",
            placeholder="np. Zarządzanie ryzykiem, Design thinking...",
        )
    with col_add2:
        st.write("")
        st.write("")
        add_custom = st.button("Dodaj do listy", key="add_custom_btn")

    if add_custom and custom_name.strip():
        if custom_name.strip() not in st.session_state["custom_competencies"]:
            st.session_state["custom_competencies"].append(custom_name.strip())
            st.rerun()

    if st.session_state["custom_competencies"]:
        st.markdown("**Własne kompetencje:**")
        to_remove = []
        for i, cn in enumerate(st.session_state["custom_competencies"]):
            col_c, col_x = st.columns([5, 1])
            col_c.write(f"+ {cn}")
            if col_x.button("Usuń", key=f"rm_custom_{i}"):
                to_remove.append(cn)
        if to_remove:
            for cn in to_remove:
                st.session_state["custom_competencies"].remove(cn)
            st.rerun()

    all_selected.extend(st.session_state.get("custom_competencies", []))
    total_count = len(all_selected)

    if total_count > 10:
        st.warning(f"Wybrano {total_count} kompetencji — max 10. Usuń nadmiarowe.")
    elif total_count > 0:
        st.info(f"Wybrano **{total_count}** kompetencji: {', '.join(all_selected)}")
    else:
        st.info("Wybierz kompetencje z katalogu lub dopisz własne.")

    st.divider()
    st.subheader("Krok 4: Generuj opisy kompetencji")

    if st.button("Generuj profil kompetencji", type="primary",
                 disabled=(not api_key or total_count == 0 or total_count > 10)):
        _curr_tasks = st.session_state.get("key_tasks_list", [])
        if not position or not _curr_tasks:
            st.warning("Podaj nazwę stanowiska i dodaj przynajmniej jedno zadanie (Krok 1–2).")
        else:
            company = CompanyProfile(
                company_name=company_name, industry=industry, size=size,
                culture_notes=culture, position_name=position, position_level=level,
                key_tasks=tasks_to_str(_curr_tasks),
                key_tasks_list=_curr_tasks,
                additional_context=extra,
            )
            with st.spinner(f"Generowanie opisów {total_count} kompetencji (30-90 sek.)..."):
                try:
                    profile = generate_competency_profile(company, all_selected, api_key)
                    st.session_state["profile"] = profile
                    st.session_state["company"] = company
                    st.success(
                        f"Wygenerowano {len(profile.competencies)} kompetencji. "
                        "Przejdź do zakładki 2 → ustaw poziomy oczekiwane i zapisz stanowisko."
                    )
                except Exception as e:
                    st.error(f"Błąd: {e}")

    if "profile" in st.session_state:
        profile: CompetencyProfile = st.session_state["profile"]
        st.divider()
        st.subheader(f"Profil: {profile.company.position_name} ({profile.company.company_name})")
        st.caption(
            f"{len(profile.competencies)} kompetencji"
            + (f" | {profile.created_at[:19]}" if profile.created_at else "")
        )

        comp_to_remove = None
        for i, comp in enumerate(profile.competencies):
            cat_display = CATEGORY_LABELS.get(comp.category, comp.category)
            with st.expander(f"**{i+1}. {comp.name}** [{cat_display}]", expanded=(i < 2)):
                st.markdown(f"**Definicja:** {comp.definition}")
                if comp.indicators:
                    st.markdown("**Wskaźniki behawioralne:**")
                    for ind in comp.indicators:
                        st.markdown(f"- {ind}")
                level_cols = st.columns(5)
                for j, lk in enumerate(LEVEL_KEYS):
                    with level_cols[j]:
                        desc = comp.get_level_description(lk)
                        color = LEVEL_COLORS[lk]
                        st.markdown(
                            f"<div style='background:{color}; color:white; padding:8px; "
                            f"border-radius:6px; font-size:0.8em; min-height:120px;'>"
                            f"<b>{lk}. {LEVEL_LABELS[lk]}</b><br><br>{desc}</div>",
                            unsafe_allow_html=True,
                        )
                if st.toggle("✏️ Edytuj opisy", key=f"toggle_edit_{i}"):
                    st.text_area(
                        "Definicja",
                        value=comp.definition,
                        key=f"edit_def_{i}",
                        height=80,
                    )
                    st.text_area(
                        "Wskaźniki behawioralne (jeden na linię)",
                        value="\n".join(comp.indicators),
                        key=f"edit_ind_{i}",
                        height=100,
                    )
                    lv_cols = st.columns(5)
                    for j, lk in enumerate(LEVEL_KEYS):
                        with lv_cols[j]:
                            st.text_area(
                                f"Poziom {lk} — {LEVEL_LABELS[lk]}",
                                value=comp.get_level_description(lk),
                                key=f"edit_lv_{i}_{lk}",
                                height=130,
                            )
                    if st.button("Zastosuj zmiany", key=f"apply_edit_{i}"):
                        comp.definition = st.session_state[f"edit_def_{i}"]
                        comp.indicators = [
                            x.strip() for x in
                            st.session_state[f"edit_ind_{i}"].splitlines()
                            if x.strip()
                        ]
                        for lk in LEVEL_KEYS:
                            setattr(comp, f"level_{lk}",
                                    st.session_state[f"edit_lv_{i}_{lk}"])
                        st.session_state["profile"] = profile
                        st.rerun()

                if st.button(f"Usuń kompetencję '{comp.name}'", key=f"rm_comp_{i}"):
                    comp_to_remove = i

        if comp_to_remove is not None:
            profile.competencies.pop(comp_to_remove)
            st.session_state["profile"] = profile
            st.rerun()

        st.divider()
        col_d1, col_d2 = st.columns([3, 1])
        with col_d1:
            new_comp_name = st.text_input(
                "Dopisz kompetencję do profilu (AI wygeneruje opis)",
                key="add_comp_to_profile",
                placeholder="Wpisz nazwę kompetencji...",
            )
        with col_d2:
            st.write("")
            st.write("")
            add_to_profile = st.button("Dopisz i generuj opis", key="add_comp_btn")

        if add_to_profile and new_comp_name.strip() and api_key:
            company = st.session_state.get("company", profile.company)
            with st.spinner(f"Generowanie opisu '{new_comp_name.strip()}'..."):
                try:
                    new_comp = generate_single_competency(new_comp_name.strip(), company, api_key)
                    profile.competencies.append(new_comp)
                    st.session_state["profile"] = profile
                    st.success(f"Dodano: {new_comp.name}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Błąd: {e}")

        st.divider()
        try:
            from stk_export import build_profile_docx
            _prof_docx = build_profile_docx(profile, None)
            _safe_pos = profile.company.position_name.replace(" ", "_")[:40]
            st.download_button(
                "Pobierz profil kompetencji DOCX",
                data=_prof_docx,
                file_name=f"Profil_{_safe_pos}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        except Exception as _e:
            st.caption(f"DOCX niedostępny: {_e}")


# ===========================================================================
# ZAKŁADKA 2: Analiza potrzeb — tylko poziom oczekiwany
# ===========================================================================
with tab2:
    st.header("Analiza potrzeb szkoleniowych")
    st.caption(
        "Ustaw **poziom oczekiwany** dla każdej kompetencji (1-5). "
        "Poziom aktualny zostanie określony na podstawie wyników testu STK."
    )

    if "profile" not in st.session_state:
        st.info("Najpierw wygeneruj profil kompetencji w zakładce 1.")
    else:
        profile: CompetencyProfile = st.session_state["profile"]

        assessor_name = st.text_input("Imię i nazwisko oceniającego / grupy", key="assess_name")
        assessor_role = st.selectbox(
            "Rola", ["trener", "przełożony", "zespół", "uczestnik (samoocena)"], key="assess_role"
        )
        st.divider()

        needs_items: list[NeedsAssessmentItem] = []
        for i, comp in enumerate(profile.competencies):
            col1, col2, col3 = st.columns([3, 2, 1])
            with col1:
                st.markdown(f"**{comp.name}**")
                cat_lbl = CATEGORY_LABELS.get(comp.category, comp.category)
                st.caption(cat_lbl)
            with col2:
                desired = st.select_slider(
                    "Poziom oczekiwany",
                    options=[1, 2, 3, 4, 5],
                    value=4,
                    format_func=lambda x: f"{x} — {LEVEL_LABELS[x][:20]}",
                    key=f"des_{i}",
                )
            with col3:
                importance = st.select_slider(
                    "Ważność",
                    options=[1, 2, 3, 4, 5],
                    value=3,
                    key=f"imp_{i}",
                )
            needs_items.append(NeedsAssessmentItem(
                competency_name=comp.name,
                current_level=0,   # określany przez wynik STK, nie samoocenę
                desired_level=desired,
                importance=importance,
                assessor=assessor_role,
            ))

        st.divider()

        if "needs_assessment" in st.session_state:
            assessment: NeedsAssessment = st.session_state["needs_assessment"]
            st.subheader("Podgląd wyników analizy potrzeb")
            sorted_items = sorted(assessment.items,
                                  key=lambda x: x.desired_level * x.importance, reverse=True)
            header_cols = st.columns([3, 2, 1, 1])
            for col_w, label in zip(header_cols, ["Kompetencja", "Poziom oczekiwany", "Ważność", "Priorytet"]):
                col_w.markdown(f"**{label}**")
            for item in sorted_items:
                cols = st.columns([3, 2, 1, 1])
                cols[0].write(item.competency_name)
                cols[1].write(f"{item.desired_level} — {LEVEL_LABELS[item.desired_level][:15]}")
                cols[2].write(str(item.importance))
                cols[3].write(str(item.desired_level * item.importance))
            st.divider()

        col_save1, col_save2 = st.columns(2)

        with col_save1:
            if st.button("Zapisz ocenę potrzeb (podgląd)", key="btn_preview_needs"):
                assessment = NeedsAssessment(
                    items=needs_items,
                    assessor_name=assessor_name,
                    assessor_role=assessor_role,
                )
                st.session_state["needs_assessment"] = assessment
                st.rerun()

        with col_save2:
            if st.button("Zapisz stanowisko i dodaj kolejne →", type="primary",
                         key="btn_save_position"):
                assessment = NeedsAssessment(
                    items=needs_items,
                    assessor_name=assessor_name,
                    assessor_role=assessor_role,
                )
                st.session_state["needs_assessment"] = assessment
                # Zapisz do listy stanowisk (z incydentami jeśli są)
                st.session_state["saved_positions"].append({
                    "profile": st.session_state["profile"],
                    "needs": st.session_state["needs_assessment"],
                    "incidents": list(st.session_state.get("incidents", [])),
                })
                n = len(st.session_state["saved_positions"])
                _reset_form()
                st.success(f"Stanowisko #{n} zapisane! Formularz zresetowany — możesz dodać kolejne.")
                st.rerun()

        st.info(
            "Po zapisaniu stanowiska formularz się zresetuje — możesz dodać kolejne.\n\n"
            "Gotowe pliki do pobrania pojawią się w **panelu bocznym** po lewej."
        )


# ===========================================================================
# ZAKŁADKA 3: Incydenty krytyczne
# ===========================================================================
with tab3:
    st.header("Incydenty krytyczne")
    st.markdown(
        "Opisz konkretne sytuacje z pracy powiązane z wybranymi kompetencjami. "
        "Na ich podstawie trener stworzy bardziej realistyczne i trafne pytania testu STK."
    )

    with st.expander("📖 Instrukcja — jak opisać incydent krytyczny", expanded=False):
        st.markdown(
            """
**Incydent krytyczny** to konkretna sytuacja z Twojej pracy, która:
- wymagała podjęcia świadomej decyzji lub działania,
- miała pozytywne lub negatywne konsekwencje,
- jest związana z konkretną kompetencją.

**Nie musi to być sytuacja wyjątkowa** — codzienne sytuacje zawodowe są równie wartościowe.

#### Jak opisać incydent — 7 elementów:
1. **Sytuacja** — Co się wydarzyło? Kiedy, gdzie, w jakim kontekście?
2. **Osoby zaangażowane** — Kto brał udział? (imion nie podawaj — np. "przełożony", "klient", "współpracownik")
3. **Decyzja / działanie** — Co zostało zrobione lub zdecydowane?
4. **Powód** — Dlaczego podjęto taką decyzję?
5. **Rezultat** — Jaki był wynik? Pozytywny czy negatywny?
6. **Najlepsza reakcja** — Jak powinien zachować się ktoś z najwyższym poziomem tej kompetencji?
7. **Najgorsza reakcja** — Jak mógłby zachować się ktoś bez tej kompetencji?

> *Źródło metodyki: Prokopowicz, Żmuda, Król (2014). Kompetencyjne testy sytuacyjne. Wolters Kluwer.*
            """
        )

    if "profile" not in st.session_state:
        st.info("Najpierw wygeneruj profil kompetencji w zakładce 1.")
    else:
        profile: CompetencyProfile = st.session_state["profile"]
        comp_names = [c.name for c in profile.competencies]

        st.subheader("Dodaj nowy incydent")
        col_a, col_b = st.columns([2, 1])
        with col_a:
            inc_comp = st.selectbox(
                "Kompetencja, której dotyczy incydent",
                comp_names,
                key="inc_comp_new",
            )
        with col_b:
            st.write("")

        inc_situation = st.text_area(
            "1. Sytuacja — co się wydarzyło? kiedy, gdzie, kontekst?",
            key="inc_sit_new", height=90,
            placeholder="Np. Podczas cotygodniowego spotkania zespołu jeden ze współpracowników..."
        )
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            inc_actors = st.text_input(
                "2. Osoby zaangażowane (bez imion)",
                key="inc_actors_new",
                placeholder="Np. przełożony, dwóch współpracowników, klient zewnętrzny"
            )
        with col_c2:
            pass

        inc_action = st.text_area(
            "3. Decyzja / działanie — co zostało zrobione?",
            key="inc_act_new", height=80,
            placeholder="Np. Zdecydowałem się porozmawiać bezpośrednio z..."
        )
        inc_reasoning = st.text_area(
            "4. Powód decyzji — dlaczego tak?",
            key="inc_reas_new", height=60,
            placeholder="Np. Uznałem, że bezpośrednia rozmowa jest skuteczniejsza niż..."
        )
        inc_result = st.text_area(
            "5. Rezultat — co się stało po tej decyzji?",
            key="inc_res_new", height=60,
            placeholder="Np. Sytuacja się uspokoiła / Konfliktu nie udało się rozwiązać..."
        )

        st.divider()
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            inc_best = st.text_area(
                "6. Najlepsza możliwa reakcja w tej sytuacji",
                key="inc_best_new", height=80,
                placeholder="Np. Osoba z wysoką kompetencją powinna najpierw..."
            )
        with col_d2:
            inc_worst = st.text_area(
                "7. Najgorsza możliwa reakcja w tej sytuacji",
                key="inc_worst_new", height=80,
                placeholder="Np. Ktoś bez tej kompetencji mógłby zignorować sytuację i..."
            )

        if st.button("➕ Dodaj incydent", type="primary", key="btn_add_incident"):
            if not inc_situation.strip():
                st.warning("Opisz sytuację (pole 1) — to minimalny wymagany element incydentu.")
            else:
                new_inc = CriticalIncident(
                    competency_name=inc_comp,
                    situation=inc_situation.strip(),
                    actors=inc_actors.strip(),
                    action=inc_action.strip(),
                    reasoning=inc_reasoning.strip(),
                    result=inc_result.strip(),
                    best_alternative=inc_best.strip(),
                    worst_alternative=inc_worst.strip(),
                )
                st.session_state["incidents"].append(new_inc)
                # Wyczyść pola formularza
                for key in ["inc_comp_new", "inc_sit_new", "inc_actors_new", "inc_act_new",
                            "inc_reas_new", "inc_res_new", "inc_best_new", "inc_worst_new"]:
                    st.session_state.pop(key, None)
                st.success(f"Dodano incydent #{len(st.session_state['incidents'])} dla kompetencji: {inc_comp}")
                st.rerun()

        # --- Lista zapisanych incydentów ---
        incidents: list[CriticalIncident] = st.session_state["incidents"]
        if incidents:
            st.divider()
            st.subheader(f"Zapisane incydenty ({len(incidents)})")

            inc_to_del = None
            for idx, inc in enumerate(incidents):
                with st.expander(
                    f"{idx+1}. [{inc.competency_name}]  {inc.situation[:70]}...",
                    expanded=False
                ):
                    st.markdown(f"**Sytuacja:** {inc.situation}")
                    if inc.actors:
                        st.markdown(f"**Osoby:** {inc.actors}")
                    if inc.action:
                        st.markdown(f"**Decyzja:** {inc.action}")
                    if inc.reasoning:
                        st.markdown(f"**Powód:** {inc.reasoning}")
                    if inc.result:
                        st.markdown(f"**Rezultat:** {inc.result}")
                    if inc.best_alternative:
                        st.markdown(f"**Najlepsza reakcja:** {inc.best_alternative}")
                    if inc.worst_alternative:
                        st.markdown(f"**Najgorsza reakcja:** {inc.worst_alternative}")
                    if st.button(f"Usuń incydent #{idx+1}", key=f"del_inc_{idx}"):
                        inc_to_del = idx

            if inc_to_del is not None:
                st.session_state["incidents"].pop(inc_to_del)
                st.rerun()

            st.divider()
            st.subheader("Pobierz incydenty")

            safe_pos = (
                profile.company.position_name.replace(" ", "_")
                if "profile" in st.session_state else "warsztat"
            )
            try:
                from stk_export import build_incidents_docx, build_incidents_xlsx
                col_e1, col_e2 = st.columns(2)
                with col_e1:
                    docx_bytes = build_incidents_docx(
                        incidents,
                        position_name=profile.company.position_name,
                        company_name=profile.company.company_name,
                    )
                    st.download_button(
                        "Pobierz DOCX (raport dla trenera)",
                        data=docx_bytes,
                        file_name=f"Incydenty_{safe_pos}.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                    )
                with col_e2:
                    xlsx_bytes = build_incidents_xlsx(
                        incidents,
                        position_name=profile.company.position_name,
                        company_name=profile.company.company_name,
                    )
                    st.download_button(
                        "Pobierz XLSX (arkusz)",
                        data=xlsx_bytes,
                        file_name=f"Incydenty_{safe_pos}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            except Exception as e:
                st.caption(f"Eksport niedostępny: {e}")

            st.info(
                "Incydenty zostaną automatycznie dołączone do JSON gdy klikniesz "
                "**'Zapisz stanowisko i dodaj kolejne'** w zakładce 2."
            )
        else:
            st.info(
                "Nie dodano jeszcze żadnych incydentów. "
                "Opisz 2–5 sytuacji per kompetencja — im więcej, tym trafniejszy test STK."
            )


# ===========================================================================
# TAB OSP: Opis Stanowiska Pracy
# ===========================================================================
with tab_osp:
    st.header("Opis Stanowiska Pracy")
    st.caption("Formularz OSP RES 1.4 — Enterprise Advisors. Wypełnij dowolną liczbę sekcji.")

    _osp_prof = st.session_state.get("profile")
    _osp_needs = st.session_state.get("needs_assessment")

    if _osp_prof:
        if st.button("Wypełnij z profilu kompetencji", key="osp_fill_btn", type="primary"):
            st.session_state["_osp_nazwa"] = _osp_prof.company.position_name
            st.session_state["_osp_jednostka"] = _osp_prof.company.company_name
            if _osp_prof.company.key_tasks_list:
                _osp_tasks = [t["name"] for t in _osp_prof.company.key_tasks_list if t.get("name")][:6]
                st.session_state["_osp_zadania"] = "\n".join(_osp_tasks)
            elif _osp_prof.company.key_tasks:
                _osp_tasks = [t.strip() for t in _osp_prof.company.key_tasks.split("\n") if t.strip()][:6]
                st.session_state["_osp_zadania"] = "\n".join(_osp_tasks)
            _osp_dm: dict = {}
            if _osp_needs:
                for _ni in _osp_needs.items:
                    _osp_dm[_ni.competency_name] = _ni.desired_level
            _osp_comps = _osp_prof.competencies[:10]
            st.session_state["_osp_komp_count"] = max(len(_osp_comps), st.session_state.get("_osp_komp_count", 4))
            for _ci, _comp in enumerate(_osp_comps):
                st.session_state[f"_osp_komp_n_{_ci}"] = _comp.name
                _lv = _osp_dm.get(_comp.name, "")
                st.session_state[f"_osp_komp_p_{_ci}"] = str(_lv) if _lv else ""
            st.rerun()
    else:
        st.info("Opcjonalnie: wygeneruj profil kompetencji w zakładce 1, aby auto-uzupełnić część pól.")

    st.divider()

    with st.expander("1. Identyfikacja stanowiska", expanded=True):
        _oc1, _oc2 = st.columns([1, 3])
        with _oc1:
            st.text_input("Numer stanowiska", key="_osp_numer")
        with _oc2:
            st.text_input("Nazwa stanowiska", key="_osp_nazwa")

    with st.expander("2. Cel stanowiska", expanded=True):
        st.text_area("Cel stanowiska", key="_osp_cel", height=100,
                     placeholder="Krótki opis celu istnienia stanowiska w organizacji...",
                     label_visibility="collapsed")

    with st.expander("3. Umiejscowienie w strukturze", expanded=False):
        _oc1, _oc2 = st.columns(2)
        with _oc1:
            st.text_input("Jednostka organizacyjna", key="_osp_jednostka")
            st.text_input("Komórka wykonawcza", key="_osp_komorka")
        with _oc2:
            st.text_input("Stanowisko przełożonego", key="_osp_przelozony")
            st.text_input("Stanowiska podległe", key="_osp_podlegle")

    with st.expander("4. Doświadczenie i kwalifikacje", expanded=False):
        _oh = st.columns([2, 3, 3])
        _oh[0].markdown("**Kryterium**"); _oh[1].markdown("**Wymagane**"); _oh[2].markdown("**Pożądane**")
        for _ol, _okw, _okp in [
            ("Staż pracy", "_osp_staz_wym", "_osp_staz_poz"),
            ("Stanowisko", "_osp_stan_wym", "_osp_stan_poz"),
            ("Specjalność", "_osp_spec_wym", "_osp_spec_poz"),
            ("Wykształcenie", "_osp_wyk_wym", "_osp_wyk_poz"),
            ("Kwalifikacje", "_osp_kwal_wym", "_osp_kwal_poz"),
            ("Uprawnienia", "_osp_upr_wym", "_osp_upr_poz"),
        ]:
            _or = st.columns([2, 3, 3])
            _or[0].write(_ol)
            _or[1].text_input("w", key=_okw, label_visibility="collapsed")
            _or[2].text_input("p", key=_okp, label_visibility="collapsed")

    with st.expander("5. Upoważnienia i odpowiedzialność", expanded=False):
        st.text_area("Upoważnienia", key="_osp_upowaznienia", height=100,
                     placeholder="Zakres upoważnień i odpowiedzialności...",
                     label_visibility="collapsed")

    with st.expander("6. Kluczowe wskaźniki efektywności (KPI)", expanded=False):
        st.caption("Jeden wskaźnik na linię — prawa kolumna to odpowiadające kryterium oceny")
        _okc1, _okc2 = st.columns(2)
        with _okc1:
            st.markdown("**Wskaźnik (KPI)**")
            st.text_area("kpi_w", key="_osp_kpi_wskazniki", height=120, label_visibility="collapsed",
                         placeholder="np.\nRealizacja planu sprzedaży (%)\nSatysfakcja klienta (NPS)")
        with _okc2:
            st.markdown("**Kryterium oceny**")
            st.text_area("kpi_k", key="_osp_kpi_kryteria", height=120, label_visibility="collapsed",
                         placeholder="np.\n≥100% kwartalnie\nNPS ≥ 8")

    with st.expander("7. Zadania wykonywane na stanowisku", expanded=False):
        st.text_area("Zadania (jeden na linię)", key="_osp_zadania", height=150,
                     placeholder="np.\nZarządzanie zespołem handlowym\nRealizacja planu sprzedaży\nCoaching handlowców")

    with st.expander("8. Kompetencje", expanded=False):
        if "_osp_komp_count" not in st.session_state:
            st.session_state["_osp_komp_count"] = 4
        _okc = st.session_state["_osp_komp_count"]
        _och = st.columns([0.4, 3.5, 1.5])
        _och[0].markdown("**Nr**"); _och[1].markdown("**Kompetencja**"); _och[2].markdown("**Poziom (1-5)**")
        for _ki in range(_okc):
            _okr = st.columns([0.4, 3.5, 1.5])
            _okr[0].write(f"{_ki + 1}.")
            _okr[1].text_input("n", key=f"_osp_komp_n_{_ki}", label_visibility="collapsed")
            _okr[2].selectbox("p", ["", "1", "2", "3", "4", "5"],
                              key=f"_osp_komp_p_{_ki}", label_visibility="collapsed")
        if st.button("+ Dodaj wiersz kompetencji", key="osp_add_komp"):
            st.session_state["_osp_komp_count"] = _okc + 1
            st.rerun()

    with st.expander("9. Pozostałe elementy opisu", expanded=False):
        for _ol9, _ok9, _oph9 in [
            ("Wyposażenie i środki pracy", "_osp_wyposazenie", "np. komputer, samochód służbowy..."),
            ("Uciążliwość", "_osp_uciazkliwosc", "np. praca przy komputerze > 4h dziennie"),
            ("Zagrożenia", "_osp_zagrozenia", "np. stres, wypalenie zawodowe"),
            ("Kto zastępuje w trakcie urlopu?", "_osp_zast_urlop", ""),
            ("Inne wymogi", "_osp_inne_wymogi", ""),
        ]:
            st.text_input(_ol9, key=_ok9, placeholder=_oph9)

    with st.expander("10. Relacje i współpraca", expanded=False):
        st.text_area("Współpraca wewnętrzna (działy, funkcje)", key="_osp_wspolpr_wewn", height=80)
        st.text_area("Współpraca zewnętrzna (klienci, dostawcy, partnerzy)", key="_osp_wspolpr_zewn", height=80)

    with st.expander("11. Zastępstwa", expanded=False):
        st.text_input("Kogo zastępuje to stanowisko", key="_osp_zast_kogo")
        st.text_input("Kto zastępuje to stanowisko", key="_osp_kto_zast")

    with st.expander("12. Limity decyzyjne", expanded=False):
        st.text_area("Decyzje podejmowane samodzielnie", key="_osp_dec_sam", height=80)
        st.text_area("Decyzje wymagające akceptacji przełożonego", key="_osp_dec_akc", height=80)
        st.text_input("Limity finansowe (kwoty, budżet)", key="_osp_limity_fin")
        st.text_input("Pełnomocnictwa i upoważnienia", key="_osp_pelnomocnictwa")

    with st.expander("13. Miejsce i tryb pracy", expanded=False):
        st.text_input("Miejsce wykonywania pracy", key="_osp_miejsce")
        st.text_input("Tryb pracy (stacjonarny, hybrydowy, zdalny)", key="_osp_tryb")
        st.text_input("Mobilność i delegacje", key="_osp_mobilnosc")
        st.text_input("Dyspozycyjność (zmiany, dyżury)", key="_osp_dyspozycyjnosc")

    st.divider()
    st.subheader("Pobierz OSP")

    _osp_komp_list = []
    for _ki in range(st.session_state.get("_osp_komp_count", 4)):
        _on = st.session_state.get(f"_osp_komp_n_{_ki}", "")
        if _on:
            _osp_komp_list.append({
                "nr": str(_ki + 1),
                "nazwa": _on,
                "poziom": st.session_state.get(f"_osp_komp_p_{_ki}", ""),
            })

    _osp_export = {
        "numer": st.session_state.get("_osp_numer", ""),
        "nazwa_stanowiska": st.session_state.get("_osp_nazwa", ""),
        "cel": st.session_state.get("_osp_cel", ""),
        "jednostka": st.session_state.get("_osp_jednostka", ""),
        "komorka": st.session_state.get("_osp_komorka", ""),
        "przelozony": st.session_state.get("_osp_przelozony", ""),
        "podlegle": st.session_state.get("_osp_podlegle", ""),
        "staz_wym": st.session_state.get("_osp_staz_wym", ""),
        "staz_poz": st.session_state.get("_osp_staz_poz", ""),
        "stanowisko_wym": st.session_state.get("_osp_stan_wym", ""),
        "stanowisko_poz": st.session_state.get("_osp_stan_poz", ""),
        "specjalnosc_wym": st.session_state.get("_osp_spec_wym", ""),
        "specjalnosc_poz": st.session_state.get("_osp_spec_poz", ""),
        "wyksztalcenie_wym": st.session_state.get("_osp_wyk_wym", ""),
        "wyksztalcenie_poz": st.session_state.get("_osp_wyk_poz", ""),
        "kwalifikacje_wym": st.session_state.get("_osp_kwal_wym", ""),
        "kwalifikacje_poz": st.session_state.get("_osp_kwal_poz", ""),
        "uprawnienia_wym": st.session_state.get("_osp_upr_wym", ""),
        "uprawnienia_poz": st.session_state.get("_osp_upr_poz", ""),
        "upowaznienia": st.session_state.get("_osp_upowaznienia", ""),
        "kpi_wskazniki": st.session_state.get("_osp_kpi_wskazniki", ""),
        "kpi_kryteria": st.session_state.get("_osp_kpi_kryteria", ""),
        "zadania": st.session_state.get("_osp_zadania", ""),
        "kompetencje": _osp_komp_list,
        "wyposazenie": st.session_state.get("_osp_wyposazenie", ""),
        "uciazkliwosc": st.session_state.get("_osp_uciazkliwosc", ""),
        "zagrozenia": st.session_state.get("_osp_zagrozenia", ""),
        "zastepstwo_urlop": st.session_state.get("_osp_zast_urlop", ""),
        "inne_wymogi": st.session_state.get("_osp_inne_wymogi", ""),
        "wspolpraca_wewn": st.session_state.get("_osp_wspolpr_wewn", ""),
        "wspolpraca_zewn": st.session_state.get("_osp_wspolpr_zewn", ""),
        "zastepuje_kogo": st.session_state.get("_osp_zast_kogo", ""),
        "kto_zastepuje": st.session_state.get("_osp_kto_zast", ""),
        "decyzje_sam": st.session_state.get("_osp_dec_sam", ""),
        "decyzje_akc": st.session_state.get("_osp_dec_akc", ""),
        "limity_fin": st.session_state.get("_osp_limity_fin", ""),
        "pelnomocnictwa": st.session_state.get("_osp_pelnomocnictwa", ""),
        "miejsce": st.session_state.get("_osp_miejsce", ""),
        "tryb": st.session_state.get("_osp_tryb", ""),
        "mobilnosc": st.session_state.get("_osp_mobilnosc", ""),
        "dyspozycyjnosc": st.session_state.get("_osp_dyspozycyjnosc", ""),
    }
    _osp_safe = (_osp_export.get("nazwa_stanowiska") or "OSP").replace(" ", "_")[:40]

    _oec1, _oec2 = st.columns(2)
    with _oec1:
        try:
            from stk_export import build_osp_docx as _build_osp_docx
            _osp_docx_bytes = _build_osp_docx(_osp_export)
            st.download_button(
                "Pobierz OSP (DOCX)",
                data=_osp_docx_bytes,
                file_name=f"OSP_{_osp_safe}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                type="primary",
            )
        except Exception as _oe:
            st.error(f"Błąd DOCX: {_oe}")
    with _oec2:
        st.download_button(
            "Pobierz OSP (JSON)",
            data=json.dumps(_osp_export, ensure_ascii=False, indent=2),
            file_name=f"OSP_{_osp_safe}.json",
            mime="application/json",
        )
